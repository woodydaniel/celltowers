"""
Export half of geocoded (success) records to a two-sheet Excel file:
  Sheet 1 — Data (125,468 records)
  Sheet 2 — Column reference table
"""

import json
import math
import random
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR   = Path(__file__).resolve().parent.parent
INPUT_FILE = BASE_DIR / "downloads" / "towers_all_classified.jsonl"
OUTPUT_DIR = BASE_DIR / "downloads"
OUTPUT_DIR.mkdir(exist_ok=True)
OUTPUT_FILE = OUTPUT_DIR / "cellmapper_towers_sample.xlsx"

COLUMN_REFERENCE = [
    ("tower_id",          "310_410_684118",                "Unique tower identifier (carrier_MCC_MNC_site)"),
    ("site_id",           "684118",                        "CellMapper site ID"),
    ("latitude",          "24.5474",                       "Tower GPS latitude (WGS84)"),
    ("longitude",         "-81.7913",                      "Tower GPS longitude (WGS84)"),
    ("provider",          "AT&T",                          "Carrier / network operator"),
    ("generation",        "4G",                            "Network generation (2G / 3G / 4G / 5G)"),
    ("site_type",         "Tower",                         "Physical site type (Tower, Rooftop, Unknown, etc.)"),
    ("active",            "True",                          "Whether the tower is currently active on CellMapper"),
    ("bands",             "5, 2, 12, 66",                  "Radio frequency band numbers in use (comma-separated)"),
    ("band_labels",       "B5 (850 MHz), B12 (700 MHz A)", "Human-readable band names with frequencies"),
    ("tower_name",        "Downtown Site A",               "Tower name if available (often blank)"),
    ("tower_parent",      "",                              "Parent site name if available (often blank)"),
    ("first_seen",        "2020-10-25",                    "Date this tower was first observed on CellMapper"),
    ("last_seen",         "2025-10-15",                    "Date this tower was last observed on CellMapper"),
    ("rural",             "False",                         "True = rural/out-of-bounds (no address geocoded); False = urban"),
    ("address",           "1500 Alberta St",               "Nearest street address (from Smarty reverse geocoding)"),
    ("city",              "Key West",                      "City of nearest address"),
    ("state",             "FL",                            "US state abbreviation"),
    ("zipcode",           "33040",                         "ZIP code of nearest address"),
    ("geocode_status",    "success",                       "success = address found; pending = not geocoded (rural/skipped)"),
    ("geocode_distance",  "86",                            "Distance in metres from tower coordinates to nearest address"),
    ("geocode_accuracy",  "Rooftop",                       "Smarty accuracy level: Rooftop, Parcel, Street, Zip, etc."),
]


def flatten(value):
    """Flatten list values to comma-separated strings."""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return value


def load_geocoded_records():
    records = []
    with open(INPUT_FILE, encoding="utf-8") as fh:
        for line in fh:
            rec = json.loads(line)
            if rec.get("geocode_status") == "success":
                records.append({k: flatten(v) for k, v in rec.items()})
    return records


def style_header_row(ws, n_cols):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def style_ref_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="D6E4F0")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        is_header = row_idx == 1
        is_alt    = row_idx % 2 == 0
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if is_header:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif is_alt:
                cell.fill = alt_fill


def main():
    print("Loading geocoded records ...")
    all_records = load_geocoded_records()
    total       = len(all_records)
    half        = math.ceil(total / 2)

    # Deterministic shuffle so same records each run
    random.seed(42)
    random.shuffle(all_records)
    sample = all_records[:half]

    print(f"Total geocoded : {total:,}")
    print(f"Sample size    : {half:,} (50%)")

    df = pd.DataFrame(sample)

    # Ensure column order matches COLUMN_REFERENCE
    col_order = [c[0] for c in COLUMN_REFERENCE]
    df = df[[c for c in col_order if c in df.columns]]

    print(f"Writing {OUTPUT_FILE.name} ...")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # --- Sheet 1: Data ---
        df.to_excel(writer, sheet_name="Tower Data", index=False)
        ws_data = writer.sheets["Tower Data"]
        style_header_row(ws_data, len(df.columns))

        # Auto-size columns (cap at 50)
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if len(df) else 0
            )
            ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        ws_data.freeze_panes = "A2"

        # --- Sheet 2: Column Reference ---
        ref_df = pd.DataFrame(COLUMN_REFERENCE, columns=["Column", "Example Value", "Description"])
        ref_df.to_excel(writer, sheet_name="Column Reference", index=False)
        ws_ref = writer.sheets["Column Reference"]
        style_ref_sheet(ws_ref)
        ws_ref.column_dimensions["A"].width = 22
        ws_ref.column_dimensions["B"].width = 35
        ws_ref.column_dimensions["C"].width = 65
        for row in ws_ref.iter_rows(min_row=2):
            ws_ref.row_dimensions[row[0].row].height = 20

    size_mb = OUTPUT_FILE.stat().st_size / 1_048_576
    print(f"Done — {OUTPUT_FILE}")
    print(f"File size: {size_mb:.1f} MB")
    print(f"Sheets: 'Tower Data' ({half:,} rows) + 'Column Reference' (22 rows)")


if __name__ == "__main__":
    main()
